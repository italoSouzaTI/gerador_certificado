import os
import sys
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import textwrap
from datetime import datetime
import locale

# Configurar locale para PT-BR
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'pt_BR')
    except:
        pass

def obter_caminho_base():
    """Obtém o diretório base do executável ou script"""
    if getattr(sys, 'frozen', False):
        # Executável PyInstaller
        return os.path.dirname(sys.executable)
    else:
        # Script Python
        return os.path.dirname(os.path.abspath(__file__))

def carregar_fontes():
    """Carrega as fontes Myriad Pro"""
    base_path = obter_caminho_base()
    
    # Procurar fontes em vários locais possíveis
    locais_fontes = [
        base_path,  # Pasta do executável
        os.path.join(base_path, 'fonts'),  # Subpasta fonts
        os.path.expanduser('~/Library/Fonts'),  # macOS
        'C:\\Windows\\Fonts',  # Windows
    ]
    
    myriad_semibold = None
    myriad_bold = None
    
    # Procurar Semibold
    for local in locais_fontes:
        for nome in ['Myriad Pro Semibold Italic.ttf', 'myriad-pro-semibold-italic.otf', 
                     'Myriad Pro Semibold Italic.otf', 'MYRIADPRO-SEMIBOLDIT.OTF']:
            caminho = os.path.join(local, nome)
            if os.path.exists(caminho):
                myriad_semibold = caminho
                break
        if myriad_semibold:
            break
    
    # Procurar Bold
    for local in locais_fontes:
        for nome in ['Myriad Pro Bold Italic.ttf', 'myriad-pro-bold-italic.otf', 
                     'Myriad Pro Bold Italic.otf', 'MYRIADPRO-BOLDIT.OTF']:
            caminho = os.path.join(local, nome)
            if os.path.exists(caminho):
                myriad_bold = caminho
                break
        if myriad_bold:
            break
    
    print("\n" + "="*60)
    print("VERIFICANDO FONTES MYRIAD PRO:")
    if myriad_semibold:
        print(f"  Semibold Italic: ✓ ENCONTRADA")
        print(f"    Arquivo: {os.path.basename(myriad_semibold)}")
    else:
        print(f"  Semibold Italic: ✗ NÃO ENCONTRADA")
    
    if myriad_bold:
        print(f"  Bold Italic: ✓ ENCONTRADA")
        print(f"    Arquivo: {os.path.basename(myriad_bold)}")
    else:
        print(f"  Bold Italic: ✗ NÃO ENCONTRADA")
    
    if not myriad_semibold or not myriad_bold:
        print("\n⚠️  Coloque as fontes Myriad Pro na mesma pasta do executável!")
    print("="*60 + "\n")
    
    try:
        font_frase1_3 = ImageFont.truetype(myriad_semibold if myriad_semibold else myriad_bold, 49)
        font_frase2 = ImageFont.truetype(myriad_bold if myriad_bold else myriad_semibold, 84)
        font_bold_numeration = ImageFont.truetype(myriad_semibold if myriad_semibold else myriad_bold, 39)
        font_undergrad_bold = ImageFont.truetype(myriad_bold if myriad_bold else myriad_semibold, 52)
        font_data = ImageFont.truetype(myriad_semibold if myriad_semibold else myriad_bold, 49)
        print("✓ Fontes carregadas com sucesso!\n")
        return font_frase1_3, font_frase2, font_bold_numeration, font_undergrad_bold, font_data
    except Exception as e:
        print(f"⚠ Erro ao carregar fontes: {e}")
        print("Usando fontes padrão do sistema...\n")
        try:
            font_frase1_3 = ImageFont.truetype('arial.ttf', 49)
            font_frase2 = ImageFont.truetype('arialbd.ttf', 84)
            font_bold_numeration = ImageFont.truetype('arial.ttf', 39)
            font_undergrad_bold = ImageFont.truetype('arialbd.ttf', 52)
            font_data = ImageFont.truetype('arial.ttf', 49)
            return font_frase1_3, font_frase2, font_bold_numeration, font_undergrad_bold, font_data
        except:
            font = ImageFont.load_default()
            return font, font, font, font, font

def ler_excel(caminho_excel, aba='2025'):
    """Lê os dados do arquivo Excel"""
    try:
        workbook = openpyxl.load_workbook(caminho_excel)
        
        # Tentar abrir a aba especificada ou a primeira disponível
        if aba in workbook.sheetnames:
            sheet = workbook[aba]
        else:
            print(f"⚠ Aba '{aba}' não encontrada. Usando a primeira aba disponível.")
            sheet = workbook.active
        
        dados = []
        for indice, linha in enumerate(sheet.iter_rows(min_row=2)):  # Pula cabeçalho
            nome = linha[0].value if len(linha) > 0 else None
            
            # Pular linhas vazias
            if not nome:
                continue
            
            registro = linha[1].value if len(linha) > 1 else ''
            graduacao = linha[2].value if len(linha) > 2 else ''
            key_user = linha[5].value if len(linha) > 5 and linha[5].value else None
            
            dados.append({
                'nome': str(nome).strip(),
                'registro': str(registro) if registro else '',
                'graduacao': str(graduacao) if graduacao else '',
                'key_user': key_user
            })
        
        workbook.close()
        return dados
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return []

def gerar_certificado(dados, indice, image_base, fonts, pasta_saida):
    """Gera um certificado individual"""
    font_frase1_3, font_frase2, font_bold_numeration, font_undergrad_bold, font_data = fonts
    
    name = dados['nome']
    register = dados['registro']
    undergraduation = dados['graduacao']
    keyUser = dados['key_user']
    
    # Data atual
    data_atual = datetime.now()
    try:
        data_certification = data_atual.strftime('%d de %B de %Y').upper()
    except:
        meses = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 
                 'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
        data_certification = f'{data_atual.day} DE {meses[data_atual.month-1]} DE {data_atual.year}'
    
    # Gerar numeração
    if keyUser and '/' in str(keyUser):
        numberUser, data = str(keyUser).split('/')
        keyUserNew = f'{data}/{numberUser}'
    else:
        keyUserNew = f'{indice+1}/2025'
    
    numeration = f'Numeração: {keyUserNew}'
    frase1 = f'A Associação Deo de Jiu-Jitsu, liderada pelo Grande Mestre Deoclécio Paulo - Faixa Vermelha 9º Grau, outorga a'
    frase2 = f'{name} ({register if register != name else " "}),'
    frase3_antes = f'na qualidade de faixa '
    frase3_depois = f', o presente diploma pelo grau de Mérito e Reconhecimento.'
    dataEmissao = f'BRASÍLIA - DF, {data_certification}'
    
    # Criar imagem
    image = image_base.copy()
    desenhar = ImageDraw.Draw(image)
    
    # FRASE 1 (centralizada)
    wrapped_text1 = textwrap.fill(frase1, width=110)
    y_position1 = 1210
    text_bbox1 = desenhar.textbbox((0, 0), wrapped_text1, font=font_frase1_3)
    text_width1 = text_bbox1[2] - text_bbox1[0]
    x_position1 = (image.width - text_width1) / 2
    desenhar.text((x_position1, y_position1), wrapped_text1, fill='black', font=font_frase1_3, align="center", spacing=60)
    
    # FRASE 2 (Nome - centralizado e bold)
    wrapped_text2 = textwrap.fill(frase2, width=110)
    y_position2 = 1310
    text_bbox2 = desenhar.textbbox((0, 0), wrapped_text2, font=font_frase2)
    text_width2 = text_bbox2[2] - text_bbox2[0]
    x_position2 = (image.width - text_width2) / 2
    desenhar.text((x_position2, y_position2), wrapped_text2, fill='black', font=font_frase2, align="center", spacing=60)
    
    # FRASE 3 (com graduação em negrito)
    y_position3 = 1430
    frase3_completa = f'{frase3_antes}{undergraduation}{frase3_depois}'
    wrapped_text3 = textwrap.fill(frase3_completa, width=110)
    
    text_bbox3 = desenhar.textbbox((0, 0), wrapped_text3, font=font_frase1_3)
    text_width3 = text_bbox3[2] - text_bbox3[0]
    x_position3 = (image.width - text_width3) / 2
    
    x_atual = x_position3
    
    # Parte 1
    desenhar.text((x_atual, y_position3), frase3_antes, fill='black', font=font_frase1_3)
    bbox_antes = desenhar.textbbox((x_atual, y_position3), frase3_antes, font=font_frase1_3)
    x_atual = bbox_antes[2]
    
    # Parte 2: GRADUAÇÃO EM NEGRITO
    desenhar.text((x_atual, y_position3), undergraduation, fill='black', font=font_undergrad_bold)
    bbox_grad = desenhar.textbbox((x_atual, y_position3), undergraduation, font=font_undergrad_bold)
    x_atual = bbox_grad[2]
    
    # Parte 3
    desenhar.text((x_atual, y_position3), frase3_depois, fill='black', font=font_frase1_3)
    
    # Numeração e data
    desenhar.text((2650, 540), numeration, fill='red', font=font_bold_numeration)
    desenhar.text((1350, 1580), dataEmissao, fill='black', font=font_data)
    
    # Salvar como PDF
    filename_pdf = os.path.join(pasta_saida, f'{indice + 1}-{name}.pdf')
    
    if image.mode != 'RGB':
        image = image.convert('RGB')
    
    image.save(filename_pdf, 'PDF', resolution=100.0, quality=95)
    print(f"✓ Certificado gerado: {os.path.basename(filename_pdf)}")
    
    return True

def main():
    """Função principal"""
    print("\n" + "="*60)
    print("GERADOR DE CERTIFICADOS - ASSOCIAÇÃO DEO DE JIU-JITSU")
    print("="*60 + "\n")
    
    base_path = obter_caminho_base()
    
    # Definir caminhos
    pasta_dados = os.path.join(base_path, 'gerar certificados_gerados')
    pasta_saida = os.path.join(base_path, 'certificado_final')
    caminho_base_img = os.path.join(base_path, 'CERTIFICADO BASE.jpg')
    caminho_excel = os.path.join(pasta_dados, 'Controle de Diplomas.xlsx')
    
    # Verificar/criar pasta de saída
    if not os.path.exists(pasta_saida):
        os.makedirs(pasta_saida)
        print(f"✓ Pasta criada: certificado_final\n")
    
    # Verificar se existe a imagem base
    if not os.path.exists(caminho_base_img):
        print(f"❌ ERRO: Arquivo 'CERTIFICADO BASE.jpg' não encontrado!")
        print(f"   Coloque o arquivo na mesma pasta do executável.\n")
        try:
            input("Pressione ENTER para sair...")
        except:
            pass
        return
    
    # Verificar se existe o Excel
    if not os.path.exists(caminho_excel):
        print(f"❌ ERRO: Arquivo Excel não encontrado!")
        print(f"   Esperado em: {caminho_excel}")
        print(f"\n   Crie a pasta 'gerar certificados_gerados' e coloque o arquivo 'Controle de Diplomas.xlsx' dentro dela.\n")
        try:
            input("Pressione ENTER para sair...")
        except:
            pass
        return
    
    # Carregar fontes
    fonts = carregar_fontes()
    
    # Carregar imagem base
    try:
        image_base = Image.open(caminho_base_img)
        print(f"✓ Imagem base carregada: CERTIFICADO BASE.jpg\n")
    except Exception as e:
        print(f"❌ Erro ao carregar imagem base: {e}\n")
        try:
            input("Pressione ENTER para sair...")
        except:
            pass
        return
    
    # Ler dados do Excel
    print(f"Lendo dados do Excel...\n")
    dados_lista = ler_excel(caminho_excel, aba='2025')
    
    if not dados_lista:
        print("❌ Nenhum dado encontrado no Excel!\n")
        print("   Verifique se a planilha tem dados na aba '2025' ou se há outra aba com dados.\n")
        try:
            input("Pressione ENTER para sair...")
        except:
            pass
        return
    
    print(f"✓ {len(dados_lista)} registro(s) encontrado(s)\n")
    print("="*60)
    print("GERANDO CERTIFICADOS...")
    print("="*60 + "\n")
    
    # Gerar certificados
    certificados_gerados = 0
    for indice, dados in enumerate(dados_lista):
        try:
            print(f"Nome: {dados['nome']}, Graduação: {dados['graduacao']}")
            if gerar_certificado(dados, indice, image_base, fonts, pasta_saida):
                certificados_gerados += 1
            print()
        except Exception as e:
            print(f"❌ Erro ao gerar certificado para {dados['nome']}: {e}\n")
    
    # Resumo final
    print("="*60)
    print(f"TOTAL DE CERTIFICADOS GERADOS: {certificados_gerados}")
    print("="*60 + "\n")
    print(f"✓ Certificados salvos em: {pasta_saida}\n")
    
    try:
        input("Pressione ENTER para sair...")
    except:
        pass  # Ignora erro em modo não-interativo

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"\n❌ ERRO CRÍTICO: {e}\n")
        try:
            input("Pressione ENTER para sair...")
        except:
            pass

